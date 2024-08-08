#!/usr/bin/env python3

# converts type_matchups.xlsx spreadsheet into Assembly code for type_matchups.asm
# defaults to the first sheet of the workbook

from openpyxl import load_workbook

effectiveness = { 0: 'NO_EFFECT', 0.5: 'NOT_VERY_EFFECTIVE', 2: 'SUPER_EFFECTIVE' }

max_width = len('PSYCHIC_TYPE') # aesthetics matter

def format_type_name(raw:str) -> str:
    res = raw.strip().upper()
    return 'PSYCHIC_TYPE' if res == 'PSYCHIC' else res # 'PSYCHIC' is coded as 'PSYCHIC_TYPE'

if __name__ == '__main__':
    interactions = []
    wb = load_workbook('type_matchups.xlsx')
    ws = wb.active
    defenders = [format_type_name(cell.value) for cell in ws[1] if cell.value is not None]
    for i in range(len(defenders)):
        attacker = defenders[i]
        a_pad = max_width - len(attacker)
        for i, cell in enumerate(ws[2+i]):
            if cell.value not in effectiveness: continue
            effect = effectiveness[cell.value]
            defender = defenders[i-1]
            d_pad = max_width - len(defender)
            code = f'\tdb {attacker}, {" "*a_pad}{defender}, {" "*d_pad}{effect}\n'
            interactions.append(code)
    with open('./data/types/type_matchups.asm', 'w') as file:
        file.write('TypeEffects:\n')
        file.write('\t;  attacker,     defender,     *=\n')
        file.writelines(interactions)
        file.write('\tdb -1 ; end\n')